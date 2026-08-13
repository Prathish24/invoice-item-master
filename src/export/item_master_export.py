import csv
import json
from io import BytesIO, StringIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ============================================================
# FIXED INVOICE / VENDOR / CUSTOMER / SHIP-TO COLUMNS
# ============================================================

FIXED_COLUMNS = [

    # --------------------------------------------------------
    # Invoice
    # --------------------------------------------------------

    (
        "Invoice Number",
        "invoice_invoice_number",
    ),

    (
        "Invoice Date",
        "invoice_invoice_date",
    ),

    (
        "Due Date",
        "invoice_due_date",
    ),

    (
        "PO Number",
        "invoice_purchase_order_number",
    ),

    # --------------------------------------------------------
    # Vendor
    # --------------------------------------------------------

    (
        "Vendor Name",
        "invoice_vendor_name",
    ),

    (
        "Vendor Address",
        "invoice_vendor_address",
    ),

    (
        "Vendor Phone",
        "invoice_vendor_phone",
    ),

    (
        "Vendor Email",
        "invoice_vendor_email",
    ),

    # --------------------------------------------------------
    # Bill To / Customer
    # --------------------------------------------------------

    (
        "Customer Name",
        "invoice_customer_name",
    ),

    (
        "Customer Address",
        "invoice_customer_address",
    ),

    # --------------------------------------------------------
    # Ship To
    # --------------------------------------------------------

    (
        "Ship To Name",
        "invoice_ship_to_name",
    ),

    (
        "Ship To Address",
        "invoice_ship_to_address",
    ),
]


# ============================================================
# ADDITIONAL INVOICE INFORMATION
# ============================================================
#
# These are NOT forced into every CSV.
#
# They appear only when at least one invoice in the export
# actually contains the corresponding value.
#
# IMPORTANT:
# The field names here must match the fields produced by
# invoice_extractor.py and later passed through get_export_rows().
# ============================================================

ADDITIONAL_INVOICE_COLUMNS = [

    (
        "Sales Order Number",
        "invoice_sales_order_number",
    ),

    (
        "Quote Number",
        "invoice_quote_number",
    ),

    (
        "Order Date",
        "invoice_order_date",
    ),

    (
        "Ship Date",
        "invoice_ship_date",
    ),

    (
        "Delivery Date",
        "invoice_delivery_date",
    ),

    (
        "Packing Slip Number",
        "invoice_packing_slip_number",
    ),

    (
        "Customer Account Number",
        "invoice_customer_account_number",
    ),

    (
        "Vendor Account Number",
        "invoice_vendor_account_number",
    ),

    (
        "Job Number",
        "invoice_job_number",
    ),

    (
        "Project Number",
        "invoice_project_number",
    ),

    (
        "Terms",
        "invoice_terms",
    ),

    (
        "Currency",
        "invoice_currency",
    ),

    (
        "Freight",
        "invoice_freight",
    ),

    (
        "Discount",
        "invoice_discount",
    ),

    (
        "Tracking Number",
        "invoice_tracking_number",
    ),

    (
        "Salesperson",
        "invoice_salesperson",
    ),

    (
        "Tax ID",
        "invoice_tax_id",
    ),
]


# ============================================================
# ITEM MASTER COLUMNS
# ============================================================

ITEM_MASTER_COLUMNS = [

    (
        "Manufacturer Part Number",
        "manufacturer_part_number",
    ),

    (
        "Vendor Part Number",
        "vendor_part_number",
    ),

    (
        "Description",
        "description",
    ),

    (
        "UOM",
        "uom",
    ),

    (
        "Qty Ship",
        "quantity_shipped",
    ),

    (
        "Unit Price USD",
        "unit_price_usd",
    ),

    (
        "Extended Price USD",
        "extended_price_usd",
    ),
]


# ============================================================
# ACTIVE COLUMNS
# ============================================================

def _active_columns(
    export_rows: list[dict[str, Any]],
) -> list[tuple[str, str]]:
    """
    Return the columns used for this export.

    Fixed Invoice / Vendor / Customer / Ship-To columns are
    always present.

    Additional invoice columns are included only when at least
    one exported row contains a real value.

    Standard line-item columns are included when at least one
    exported line item contains a value.
    """

    # --------------------------------------------------------
    # Additional invoice-level columns
    # --------------------------------------------------------

    active_additional_columns = [

        (
            header,
            field,
        )

        for header, field
        in ADDITIONAL_INVOICE_COLUMNS

        if any(
            row.get(field)
            not in (None, "")
            for row in export_rows
        )
    ]

    # --------------------------------------------------------
    # Standard line-item columns
    # --------------------------------------------------------

    active_item_columns = [

        (
            header,
            field,
        )

        for header, field
        in ITEM_MASTER_COLUMNS

        if any(
            row.get(field)
            not in (None, "")
            for row in export_rows
        )
    ]

    # --------------------------------------------------------
    # Never allow the line-item section to be empty.
    # --------------------------------------------------------

    active_item_columns = (
        active_item_columns
        or ITEM_MASTER_COLUMNS
    )

    # --------------------------------------------------------
    # Dynamic additional information
    #
    # This is a single stable export column. It combines:
    #   - invoice_additional_info
    #   - line-item additional_info
    #
    # Different invoices can therefore carry different extra fields
    # without changing the CSV schema for every supplier.
    # --------------------------------------------------------

    has_dynamic_additional_info = any(
        isinstance(
            row.get("invoice_additional_info"),
            dict,
        )
        and bool(
            row.get("invoice_additional_info")
        )
        or isinstance(
            row.get("additional_info"),
            dict,
        )
        and bool(
            row.get("additional_info")
        )
        for row in export_rows
    )

    dynamic_columns = (
        [
            (
                "Additional Information",
                "additional_information",
            )
        ]
        if has_dynamic_additional_info
        else []
    )

    return (
        FIXED_COLUMNS
        + active_additional_columns
        + dynamic_columns
        + active_item_columns
    )


# ============================================================
# DYNAMIC ADDITIONAL INFORMATION
# ============================================================

def _build_additional_information(
    row: dict[str, Any],
) -> str:
    """
    Combine invoice-level and line-item dynamic information
    into one stable export value.

    The source dictionaries are preserved as JSON text so keys
    and values are not silently lost.
    """

    combined: dict[str, Any] = {}

    invoice_info = row.get(
        "invoice_additional_info"
    )

    if isinstance(
        invoice_info,
        dict,
    ):

        for key, value in invoice_info.items():

            if key is None:
                continue

            key = str(
                key
            ).strip()

            if not key:
                continue

            if value in (
                None,
                "",
            ):
                continue

            combined[
                f"invoice:{key}"
            ] = value

    line_info = row.get(
        "additional_info"
    )

    if isinstance(
        line_info,
        dict,
    ):

        for key, value in line_info.items():

            if key is None:
                continue

            key = str(
                key
            ).strip()

            if not key:
                continue

            if value in (
                None,
                "",
            ):
                continue

            combined[
                f"line:{key}"
            ] = value

    if not combined:
        return ""

    return json.dumps(
        combined,
        ensure_ascii=False,
        default=str,
    )


# ============================================================
# BUILD EXCEL WORKBOOK
# ============================================================

def build_item_master_workbook(
    export_rows: list[dict[str, Any]],
) -> bytes:
    """
    Build an Item Master Excel workbook.

    One row = one invoice line item.

    Every row contains:

        Invoice information
        Additional invoice information, when available
        Vendor information
        Customer / Bill-To information
        Ship-To information
        Item information
    """

    columns = _active_columns(
        export_rows
    )

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Item Master"

    # --------------------------------------------------------
    # Header styling
    # --------------------------------------------------------

    header_font = Font(
        bold=True
    )

    header_alignment = Alignment(
        horizontal="center"
    )

    # --------------------------------------------------------
    # Header row
    # --------------------------------------------------------

    for col_index, (
        header,
        _field,
    ) in enumerate(
        columns,
        start=1,
    ):

        cell = sheet.cell(
            row=1,
            column=col_index,
            value=header,
        )

        cell.font = header_font

        cell.alignment = (
            header_alignment
        )

    # --------------------------------------------------------
    # Data rows
    # --------------------------------------------------------

    for row_index, item in enumerate(
        export_rows,
        start=2,
    ):

        item = dict(
            item
        )

        item[
            "additional_information"
        ] = _build_additional_information(
            item
        )

        for col_index, (
            _header,
            field,
        ) in enumerate(
            columns,
            start=1,
        ):

            value = item.get(
                field
            )

            sheet.cell(
                row=row_index,
                column=col_index,
                value=value,
            )

    # --------------------------------------------------------
    # Column widths
    # --------------------------------------------------------

    for col_index, (
        header,
        _field,
    ) in enumerate(
        columns,
        start=1,
    ):

        sheet.column_dimensions[
            get_column_letter(
                col_index
            )
        ].width = max(
            18,
            len(header) + 2,
        )

    # --------------------------------------------------------
    # Freeze header
    # --------------------------------------------------------

    sheet.freeze_panes = "A2"

    # --------------------------------------------------------
    # Serialize workbook
    # --------------------------------------------------------

    buffer = BytesIO()

    workbook.save(
        buffer
    )

    buffer.seek(0)

    return buffer.getvalue()


# ============================================================
# BUILD CSV
# ============================================================

def build_item_master_csv(
    export_rows: list[dict[str, Any]],
) -> bytes:
    """
    Build an Item Master CSV.

    One row = one invoice line item.

    Invoice / Vendor / Customer / Ship-To information and any
    additional invoice information are repeated on every line
    so every item can be traced back to its source invoice.

    Additional columns appear only when at least one exported
    row contains a value for that field.
    """

    columns = _active_columns(
        export_rows
    )

    buffer = StringIO(
        newline=""
    )

    writer = csv.writer(
        buffer
    )

    # --------------------------------------------------------
    # Header
    # --------------------------------------------------------

    writer.writerow(
        [
            header
            for header, _field
            in columns
        ]
    )

    # --------------------------------------------------------
    # Rows
    # --------------------------------------------------------

    for raw_item in export_rows:

        item = dict(
            raw_item
        )

        item[
            "additional_information"
        ] = _build_additional_information(
            item
        )

        writer.writerow(
            [
                (
                    item.get(field)
                    if item.get(field)
                    is not None
                    else ""
                )

                for _header, field
                in columns
            ]
        )

    # --------------------------------------------------------
    # UTF-8 CSV
    # --------------------------------------------------------

    return buffer.getvalue().encode(
        "utf-8-sig"
    )
